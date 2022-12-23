# -*- coding: utf-8 -*-
"""
Created on Wed Dec 21 11:09:16 2022

@author: Gebruiker
"""

import streamlit as st
import toml
import pandas as pd
import plotly.express as px
from  PIL import Image
import io
import numpy as np
import math 
import plotly.figure_factory as ff
from PIL import Image
from streamlit_player import st_player
import matplotlib.pyplot as plt
import matplotlib.dates as md
import numpy as np
from datetime import datetime
from pylab import rc
from st_aggrid import AgGrid
import plotly

import pandas as pd
import matplotlib.pyplot as plt
from pylab import rc
import random as random
import time
import xlsxwriter
import os
from openpyxl import load_workbook
import tabulate
from tabulate import tabulate
from openpyxl.styles import PatternFill



st.set_page_config(page_title="Project 5 tool")

st.title("Checking a circulation planning")
st.sidebar.title("Background information")
st.markdown("This tool checks the optimization of a circulation planning of Transdev.")

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(["The tool","Overview","Route", "VDL and Hermes", "Transdev","User manual"])


# sidebar ---------------------------------------------------------------------------------------------------------------------------------------------
with st.sidebar:
    st.write("Remaining information explained.")
    with st.expander("About this tool"):
            st.write('Following a guest lecture, we as a group were'
                     ' given an assignment to create a tool for a '
                     ' circulation planning, which was created '
                     ' to represent the schedule of what a bus runs'
                     ' on a single day. This tool is used to'
                     ' check whether it meets certain requirements.')
    with st.expander("What is a circulation planning?"):
        st.write('A circulation planning is a schedule of a one-day'
                 ' period for one or more buses. It is what a bus'
                 ' runs in one day. For example, a bus can change'
                 ' its driver and line number in a day. Not only does'
                 ' a circulation planning show when the bus has a' 
                 ' service run, but also when the bus has a' 
                 'material run, needs to recharge, go to the garage'
                 'or when a bus is stationary (_idle_ ).')
    with st.expander("What is Transdev?"):
        st.write('Transdev is an international public transport company based in France.'
                 ' It provides 11 million passenger journeys worldwide every day.'
                 ' Transdev operates in 18 countries across five continents with some 85,000 employees.'
                 ' In the Netherlands, Transdev is the market leader in public transport, taxi and mobile care.'
                 ' It has several brands under it such as Hermes, Witte Kruis, Connexxion and ParkShuttle.'
                 ' Transdev has been innovating transport for many years. In 1999'
                 ' in fact, they already introduced the Parkshuttle, a fully autonomous electric shuttle.'
                 ' (Source: https://www.wijzijntransdev.nl/nl/over-ons/maak-kennis-met-transdev )')
    with st.expander("Zero-emission"):
        st.write('The buses on these lines are Zero Emission buses. This means that the drive'
                 ' causes no harmful emissions. The buses on these lines are'
                 ' battery buses, they run on electricity.'
                 " Transdev's spearhead is to use Zero Emission buses to" 
                 ' improve air quality, reduce noise and increase driving comfort for passengers'
                 ' and drivers.')
    with st.expander("Which bus lines?"):
            st.write('This tool only looks at circulation plannings' 
                     ' of bus lines 400 and 401.' 
                     ' Both go from Eindhoven Station'
                     ' to Eindhoven Airport and vice versa.' 
                     ' However, bus line 400 passes only two other bus stops,'
                     ' while bus line 401 passes no less than 15 other bus stops.'
                     ' Bus line 400 goes north to its final destination and' 
                     ' bus line 401 goes west. _(In the year 2022)_')

# tabblad 1 ----------------------------------------------------------------------------------------------------------------------------------            
# alle code van alle penalties met goede variabele namen

with tab1:
    count = 0
    connexxion_file = st.file_uploader("_Upload the connexxion data:_", type=['xlsx'])
    circulation_file = st.file_uploader("_Upload the circulation planning file here:_", type=['xlsx'],key=count)
    #connexxion_file = st.file_uploader("_Upload the connexxion data:_", type=['xlsx'])
    count += 1
    if circulation_file and connexxion_file is not None:
        st.balloons()


    if circulation_file:
        if connexxion_file:
            with st.spinner("Please wait..."):
                data_planning = pd.read_excel(circulation_file)
                startlocatie = data_planning.startlocatie
                eindlocatie = data_planning.eindlocatie
                starttijd = data_planning.starttijd
                eindtijd = data_planning.eindtijd
                activiteit = data_planning.activiteit
                buslijn = data_planning.buslijn
                buslijn = buslijn.fillna(0)
                omloopnummer = data_planning['omloop nummer']
                
                data = pd.read_excel(connexxion_file, sheet_name='Dienstregeling')
                Startlocatie = data.startlocatie
                Vertrektijd = data.vertrektijd
                Eindlocatie = data.eindlocatie
                Buslijn_ = data.buslijn
                
                data_afstand = pd.read_excel(connexxion_file, sheet_name='Afstand matrix')
                startlocatie_a = data_afstand.startlocatie
                eindlocatie_a = data_afstand.eindlocatie
                minimale_tijd = data_afstand['min reistijd in min']
                maximale_tijd = data_afstand['max reistijd in min']
                afstand = data_afstand['afstand in meters']
                Buslijn = data_afstand['buslijn'].fillna(0)
                
              
                df = pd.read_excel(r'data_project_05 (3).xlsx')
                tijd=df['t']
                speed=df['WheelBasedVehicleSpeed']
                voltage=df['DICO3_DCLinkVoltageDriveSystem']
                current=df['DICO3_DCLINKTractionCurrent']
                massa=df.Payload
                
                
                penalty_idle = 0
                overstaptijd=[]
                index_idle=[]
                Minuten=[]
                for i in range(len(data_planning)-1):
                    a = str(data_planning['eindtijd'][i])
                    b = str(data_planning['starttijd'][i+1])
                    c = str(data_planning['starttijd'][i])
                    t1 = datetime.strptime(a, '%H:%M:%S')
                    t2 = datetime.strptime(b,'%H:%M:%S')
                    t3 = datetime.strptime(c,'%H:%M:%S')
                    delta_t=t2-t1
                    delta_t1=t1-t3
                    delta_t=str(delta_t)
                    delta_t1=str(delta_t1)
                    overstaptijd.append(delta_t)
                    if activiteit[i]=='idle':
                        if delta_t1[3]!= 'd':
                            minuten=int(delta_t1[2:4]) + 60*int(delta_t1[0])
                            index_idle.append(i)
                            penalty_idle+=minuten/2
                            Minuten.append(minuten)
                        else:
                            minuten =int(delta_t1[10:12])+60*int(delta_t1[8])
                            penalty_idle+=minuten/2
                            index_idle.append(i)
                            Minuten.append(minuten)
                    if omloopnummer[i]==omloopnummer[i+1]:
                        if delta_t[3] != 'd':
                            minuten = int(delta_t[2:4]) + 60*int(delta_t[0])
                            if minuten != 0:
                                index_idle.append(i+1)
                                penalty_idle += minuten  
                                Minuten.append(minuten)
                        else:
                            minuten = int(delta_t[10:12]) + 60*int(delta_t[8])
                            if minuten != 0:
                                index_idle.append(i+1)
                                penalty_idle += minuten / 2
                                Minuten.append(minuten)
    
    
                #Wanneer de bus aankomt op bestemming en de bus niet gelijk doorrijdt wordt staat hij 'te lang ' stil. Of wanneer de activiteit idle time is.
                #MOETEN NOG EVEN KJKEN MET ZIJN ALLEN WELKE GRENSWAARDE WE NEMEN VOOR NIET GELIJK DOORRIJDEN.
    
    
    
                            
                            
                            
                            
                            
                            
                            
                            
                penalty_mat = 0
                index_mat=[]
                for j in range(len(data_planning)):
                    if activiteit[j]=='materiaal rit':
                        penalty_mat += 1
                        index_mat.append(j)
                index_tijdsoverschreiding=[]
                penalty_tijd = 0
                for v in range(len(data_planning)):
                    juistelijn = False
                    a = str(data_planning['eindtijd'][v])
                    b = str(data_planning['starttijd'][v])
                    # beginpunt = data_planning['startlocatie'][v]
                    # eindpunt = data_planning['eindlocatie'][v]
                    # lijn = data_planning['buslijn'][v]
                    t1 = datetime.strptime(a, '%H:%M:%S')
                    t2 = datetime.strptime(b,'%H:%M:%S')
                    reistijd = t1 - t2
                    reistijd=str(reistijd)
    
                    n = 0
                    while (not juistelijn) and n <= len(data_afstand)-1:
    
                        if (data_afstand['startlocatie'][n] == data_planning['startlocatie'][v]) and (data_afstand['eindlocatie'][n] == data_planning['eindlocatie'][v]):
                        #We gaan langs de afstandenmatrix, kijken welke rij overeenkomt met de rij uit de omloopplanning waar we zitten.
    
                            #print((not isinstance(data_afstand['buslijn'][n], np.floating)) and (not isinstance(data_planning['buslijn'][v], np.floating)))
    
                            if data_afstand['buslijn'][n] == data_planning['buslijn'][v]:
                                juistelijn = True
                                #print(n)
                            #De volgende if-statement is er omdat, volgens python, LEEG niet gelijkstaat aan LEEG, en niet alle routes zijn gelinkt aan een buslijn.   
    
                            elif pd.isna(data_afstand.loc[n,'buslijn']) and pd.isna(data_planning.loc[v,'buslijn']):
                            #(not isinstance(data_afstand['buslijn'][n], np.floating)) and (not isinstance(data_planning['buslijn'][v], np.floating)):
                                juistelijn = True
                            else:
                                n +=1
                        else:
                            n += 1
                    #if n > len(data_afstand)-1:
                        # print(f"Er gaat iets mis bij index: {v}") #Hier voor debuggen, uiteindelijk telde het alleen nog idle en momenten.
                    if n <= len(data_afstand) - 1:
                        if reistijd[3] != 'd': #Dit voor het geval dat de de begintijd voor 12 's nachts is en de eindtijd erna.
                            if (int(reistijd[2:4]) < data_afstand['min reistijd in min'][n]) or (int(reistijd[2:4]) > data_afstand['max reistijd in min'][n]):
                                penalty_tijd += 1
                                index_tijdsoverschreiding.append(v)
                        else: 
                            if int(reistijd[10:12]) < data_afstand['min reistijd in min'][n] or int(reistijd[10:12]) > data_afstand['max reistijd in min'][n]:
                                penalty_tijd +=1
                                index_tijdsoverschreiding.append(v)
    
                index_niet_regeling=[]
                penalty_dienst = 0
                for h in range(len(data)):
                    ind = 0
                    indeomloop = False
                    startpunt = data['startlocatie'][h]
                    eindpunt = data['eindlocatie'][h]
                    departure = data['vertrektijd'][h]
                    lijn = data['buslijn'][h]
                    while not indeomloop and ind <= len(data_planning)-1:
    
                        if departure == data_planning['starttijd'][ind][0:5] and startpunt == data_planning['startlocatie'][ind]: 
                            if eindpunt == data_planning['eindlocatie'][ind] and lijn == data_planning['buslijn'][ind]:
                                indeomloop = True
    
                        else:
                            ind +=1
                    if ind == len(data_planning):
                        index_niet_regeling.append(h)
                        penalty_dienst += 1
    
                        
                        
                        
                index_opladen=[]
                oplaadmomenten = []
                for i in range(len(data_planning)):
                    if activiteit[i] == 'opladen':
                        oplaadmomenten.append(i)
    
                verschilintijd = 0
                penalty_opladen = 0
                for j in oplaadmomenten:
                    eindtijd_opladen = str(data_planning['eindtijd'][j])
                    begintijd_opladen = str(data_planning['starttijd'][j])
                    begintijd_a = datetime.strptime(begintijd_opladen, '%H:%M:%S')
                    eindtijd_a = datetime.strptime(eindtijd_opladen,'%H:%M:%S')
                    verschilintijd = eindtijd_a - begintijd_a
                    delta_waarde15min = datetime.strptime('0:15:00', '%H:%M:%S') - datetime.strptime('0:00:00', '%H:%M:%S')
                    if verschilintijd <= delta_waarde15min:
                        penalty_opladen += 1
                        index_opladen.append(j)
                        
                        
                penalty_circulations=1
                omlopen=[]
                omlopen.append(omloopnummer[0])
                index_omloop=[]
                for i in range(len(data_planning)-1):
                    if omloopnummer[i]!=omloopnummer[i+1]:
                        omlopen.append(omloopnummer[i+1])
                        penalty_circulations+=1
                        index_omloop.append(i+1)
                        
    
    
                vermogen_list=[]
                reistijd=[]
                tijdsverschil=[]
    
                count=0
    
                for i in range(len(data_planning)-1):
                    a = str(data_planning['starttijd'][i])
                    b = str(data_planning['eindtijd'][i])
                    t1 = datetime.strptime(a, '%H:%M:%S')
                    t2 = datetime.strptime(b,'%H:%M:%S')
                    delta_t=t2-t1
                    delta_t=str(delta_t)
                    tijdsverschil.append(delta_t)
                    if delta_t[3]!= 'd':
                        minuten=int(delta_t[2:4]) + 60*int(delta_t[0])
                        if activiteit[i]!='idle' and activiteit[i]!='opladen':
                            reistijd.append(minuten)
                        else:
                            reistijd.append(minuten)
                    else:
                        minuten =int(delta_t[10:12])+60*int(delta_t[8])
                        if activiteit[i]!='idle' and activiteit[i]!='opladen':
                            reistijd.append(minuten)
                        else:
                            reistijd.append(minuten)
                            
                DRU=0
                for i in range(len(data_planning)):
                    if buslijn[i]!=0:
                        DRU+=reistijd[i]
                DPRU=np.sum(reistijd)  
                DD=DPRU/DRU    
    
                vermogen_dienstrit=[]
                vermogen_materiaalrit=[]
                reistijd=[]
                tijdsverschil=[]
                payload_voor_dienstrit=250 # Deze waarde is zo gekozen, zodat er genoeg data wordt langst gegaan, om een goede waarde te krijgen voor het verbruik
                kwh_idle=0.01              # per dienstrit en het verbruik per materiaal rit. Dus een massa van boven 200 kg is een dienstrit en een massa van 
                for i in range(len(df)):   # onder de 200 kg is een materiaal rit. Vanuit hier gaan we de waardes berekenen.
                    if massa[i]>payload_voor_dienstrit:
                        vermogen=voltage[i]*current[i]
                        vermogen_dienstrit.append(vermogen)
                    else:
                        vermogen=voltage[i]*current[i]
                        vermogen_materiaalrit.append(vermogen)
    
                vermogen_per_seconde_dienstrit=np.mean(vermogen_dienstrit)/3600000
                vermogen_per_minuut_dienstrit=vermogen_per_seconde_dienstrit*60
                vermogen_per_seconde_materiaalrit=np.mean(vermogen_materiaalrit)/3600000
                vermogen_per_minuut_materiaalrit=vermogen_per_seconde_materiaalrit*60
                count=0
    
                for i in range(len(data_planning)):
                    a = str(data_planning['starttijd'][i])
                    b = str(data_planning['eindtijd'][i])
                    t1 = datetime.strptime(a, '%H:%M:%S')
                    t2 = datetime.strptime(b,'%H:%M:%S')
                    delta_t=t2-t1
                    delta_t=str(delta_t)
                    tijdsverschil.append(delta_t)
                    if delta_t[3]!= 'd':
                        minuten=int(delta_t[2:4]) + 60*int(delta_t[0])
                        if activiteit[i]!='idle' and activiteit[i]!='opladen':
                            reistijd.append(minuten)
                        else:
                            reistijd.append(0)
                    else:
                        minuten =int(delta_t[10:12])+60*int(delta_t[8])
                        if activiteit[i]!='idle' and activiteit[i]!='opladen':
                            reistijd.append(minuten)
                        else:
                            reistijd.append(0)
                originele_capaciteit=350
                SOH=350*0.9
                oplaadtempo_per_minuut=250
                veiligheidsmarge=0.1*SOH
                verbruik_per_rit=[]
                capaciteit_einde_rit=[]
                begin_capaciteit=SOH
                count=1
                penalty_below_safety_margin=0
                index_te_weinig_capaciteit=[]
    
                for i in range(len(reistijd)):
                    if activiteit[i]=='dienst rit':
                        kwh_per_rit=reistijd[i]*vermogen_per_minuut_dienstrit
                    else:
                        kwh_per_rit=reistijd[i]*vermogen_per_minuut_materiaalrit
                    verbruik_per_rit.append(kwh_per_rit)
    
                for i in range(len(data_planning)):
                    if activiteit[i]!='opladen' and activiteit[i]!='idle':               
                        capaciteit=begin_capaciteit-verbruik_per_rit[i]
                    else:
                        a = str(data_planning['starttijd'][i])
                        b = str(data_planning['eindtijd'][i])
                        t1 = datetime.strptime(a, '%H:%M:%S')
                        t2 = datetime.strptime(b,'%H:%M:%S')
                        delta_t=t2-t1
                        delta_t=str(delta_t)
                        if delta_t[3]!= 'd':
                            minuten=int(delta_t[2:4]) + 60*int(delta_t[0])
                        else:
                            minuten =int(delta_t[10:12])+60*int(delta_t[8])
                        if activiteit[i]=='idle':
                            capaciteit=begin_capaciteit-kwh_idle*(minuten/60)
                        else:
                            capaciteit=begin_capaciteit+(oplaadtempo_per_minuut*(minuten/60))
                    capaciteit_einde_rit.append([capaciteit,i,activiteit[i]])
                    begin_capaciteit=capaciteit
                    if omloopnummer[i]!=omloopnummer[count]:
                        begin_capaciteit=SOH
                    count+=1
                    if count==len(data_planning)-1:
                        count=i
                    if capaciteit_einde_rit[i][0]<veiligheidsmarge:
                        penalty_below_safety_margin+=1
                        index_te_weinig_capaciteit.append(i)
                        
                total_penalty_score=penalty_dienst+penalty_below_safety_margin+penalty_opladen+penalty_idle+penalty_tijd+penalty_mat
                #%%
                col1, col2 = st.columns(2)
                with col1:
                
                    data_table=[['Idletime',penalty_idle],['Material_trip',penalty_mat],['Exceed time',penalty_tijd],
                                ['Not according to service trip',penalty_dienst],  ['charging too short',penalty_opladen]]
                    col_names=['Penalty function','Points']
                    table=(tabulate(data_table, headers=col_names))
                    
                    st.write(pd.DataFrame(np.array([    ['Idletime', [penalty_idle] ],
                                                        ['Material trip',[penalty_mat] ],
                                                        ['Exceed time',[penalty_tijd] ],
                                                        ['Not according to service trip',[penalty_dienst] ],
                                                        ['Charging too short',[penalty_opladen] ]],
                                                   dtype=object ),
                                                      columns = col_names) )
                with col2:
                    total_penalty_score=penalty_dienst+penalty_below_safety_margin+penalty_opladen+penalty_idle+penalty_tijd+penalty_mat
                    DD = np.round(DD,decimals=4)
                    data_table1=[['Total_penalty_score',total_penalty_score],['DD',DD],['Count of circulations',penalty_circulations]]
                    col_names1=['','Count']
                    table1=tabulate(data_table1,headers=col_names1)
                    st.write(pd.DataFrame(np.array([    ['Total penalty score', [total_penalty_score] ],
                                                        ['DD',[DD] ],
                                                        ['Count of circulations',[penalty_circulations] ]],
                                                   dtype=object ),
                                                      columns = col_names1) )
    #%%
                # Grab Currrent Time After Running the Code
    
                wb=xlsxwriter.Workbook('Fouls.xlsx') # Hier wordt de naam van de workbook aangemaakt van excel(die wordt later nog gewijzigd)
                ws=wb.add_worksheet('Planning')
                ws.write(0,0,'startlocatie')
                ws.write(0,1,'eindlocatie')
                ws.write(0,2,'starttijd')
                ws.write(0,3,'eindtijd')
                ws.write(0,4,'activiteit')
                ws.write(0,5,'buslijn')
                ws.write(0,6,'omloop nummer')
                bold=wb.add_format({'bold':True})
                ws.write(0,7,'not according to service trip',bold)
                ws.write(0,8,'below safety margin',bold)
                ws.write(0,9,'exceed time',bold)
                ws.write(0,10,'charging too short',bold)
                ws.write(0,11,'idletime',bold)
                ws.write(0,12,'material trip',bold)
                cell_format = wb.add_format()
                ws.write(0,14,'DD (minimize)')
                ws.write(0,15,DD,bold)
                ws.write(1,14,'count of circultions (minimize)')
                ws.write(1,15,penalty_circulations,bold)
                ws.write(2,14,'total count penaltyfunctions')
                ws.write(2,15,total_penalty_score,bold)
                ws.write(4,14,'not according to service trip')   
                ws.write(5,14,'below safety margin')
                ws.write(5,15,'\u03A9')
                ws.write(6,14,'exceed time')
                ws.write(6,15,'\u03B8')

                ws.write(7,14,'charging too short')
                ws.write(7,15,'\u03BB')
                ws.write(8,14,'idletime')
                ws.write(8,15,'')
                ws.write(9,14,'material trip')
                ws.write(9,15,'\u0394')
                row=1
                tel=0
                tel2=1
                for i in range(len(data_planning)):
                    col=0
                    ws.write(row,col,str(startlocatie[i]))
                    col+=1
                    ws.write(row,col,str(eindlocatie[i]))
                    col+=1
                    ws.write(row,col,str(starttijd[i]))
                    col+=1
                    ws.write(row,col,str(eindtijd[i]))
                    col+=1
                    ws.write(row,col,str(activiteit[i]))
                    col+=1
                    ws.write(row,col,(buslijn[i]))
                    col+=1
                    ws.write(row,col,omloopnummer[i])
                    col+=1
                    if i in index_niet_regeling:
                        ws.write(tel2,col,str(Vertrektijd[i]))
                        tel2+=1
                    col+=1
                    if i in index_te_weinig_capaciteit:
                        ws.write(row,col,'\u03A9')
                    col+=1
                    if i in index_tijdsoverschreiding:
                        ws.write(row,col,'\u03B8')
                    col+=1
                    if i in index_opladen:
                        ws.write(row,col,'\u03BB')
                    col+=1   
                    if i in index_idle:
                        ws.write(row,col,Minuten[tel])
                        tel+=1
                    col+=1
                    if i in index_mat:
                        ws.write(row,col,'\u0394')
                    row+=1

                    
                wb.close()
                from openpyxl import load_workbook
                workbook=load_workbook('Fouls.xlsx')
                worksheet=workbook['Planning']
                worksheet.column_dimensions['A'].width=10
                worksheet.column_dimensions['B'].width=10
                worksheet.column_dimensions['E'].width=10.5
                worksheet.column_dimensions['G'].width=15
                worksheet.column_dimensions['I'].width=18
                worksheet.column_dimensions['J'].width=11.5
                worksheet.column_dimensions['G'].width=15
                worksheet.column_dimensions['L'].width=8
                worksheet.column_dimensions['K'].width=16.5
                worksheet.column_dimensions['G'].width=14.5
                worksheet.column_dimensions['O'].width=25
                worksheet.column_dimensions['M'].width=11.5
                worksheet.column_dimensions['H'].width=24
                workbook.save('Fouls.xlsx')
                #%%               
                from io import BytesIO
                data = pd.read_excel(r'Fouls.xlsx')
                
                def to_excel(df_fouls):
                    output = BytesIO()
                    writer = pd.ExcelWriter(output, engine='xlsxwriter')
                    df_fouls.to_excel(writer, index=False, sheet_name='Sheet1')
                    workbook = writer.book
                    worksheet = writer.sheets['Sheet1']
                    format1 = workbook.add_format({'num_format': '0.00'}) 
                    worksheet.set_column('A:A', None, format1)  
                    writer.save()
                    processed_data = output.getvalue()
                    
                    return processed_data
                df_xlsx = to_excel(data)
                
                st.download_button(label=' Download File',
                                    data= df_xlsx,
                                    file_name= 'Fouls.xlsx')              

            

# tabblad 2 ----------------------------------------------------------------------------------------------------------------------------------

with tab2:
    uploaded_file = st.file_uploader("_Upload the circulation planning file here:_", type=['xlsx'])
    
    if uploaded_file is not None:
        with st.spinner("Please wait..."):
            st.subheader("Table of the circulation planning")
            with st.expander("Generate Table", expanded=False):
                df = pd.read_excel(uploaded_file, engine='openpyxl')
                st.dataframe(df)
    
            df = pd.read_excel(uploaded_file, engine='openpyxl')
            buslijn = df['buslijn']
            starttijd = df['starttijd']
            eindtijd = df['eindtijd']
            activiteit = df['activiteit']
            tasks = df['omloop nummer']
            start = pd.to_datetime(df['starttijd'])
            finish = pd.to_datetime(df['eindtijd'])
            kleur = df['activiteit'] 
            colors = {'materiaal rit': '#FA0808','dienst rit':'#0984DE', 'idle':'#030303' ,'opladen': '#17F502'}
            st.subheader('Gantt diagram van de omloopplanning')
            with st.spinner("Please wait..."):
                with st.expander("Generate Gantt Chart", expanded=False):
                    lijst = []
                    for i in range(len(df)):
                        lijst.append(dict(Bus=str(buslijn[i]), Start=str(starttijd[i]), End=str(eindtijd[i]), Tijdsbesteding=str(activiteit[i])))
                       
                    df_lijst = pd.DataFrame(lijst)
        
                    #Omzetten naar tijdreeksen waarbij ook word gekeken naar de dagen
                    start = pd.to_datetime(df['starttijd'])
                    end = pd.to_datetime(df['eindtijd'])
        
                    for i in range(len(df)):
                        if (str(start[i])[11:13]) < '05':
                            start[i] = start.apply(lambda x: x.replace(year=2023, month=1, day=2))[i]
                        else:
                            start[i] = start.apply(lambda x: x.replace(year=2023, month=1, day=1))[i]
        
                    for i in range(len(df)):
                        if (str(end[i])[11:13]) < '05':
                            end[i] = end.apply(lambda x: x.replace(year=2023, month=1, day=2))[i]
                        else:
                            end[i] = end.apply(lambda x: x.replace(year=2023, month=1, day=1))[i]
                           
                    #plotten van het figuur
                    fig = px.timeline(df_lijst, x_start=start, x_end=end, y=tasks,color_discrete_map = colors, color=kleur, hover_name = 'Tijdsbesteding', title= 'Circulation planning')
        
                    fig.update_layout(
                                            title_font_size=35,
                                            font_size=16,
                                            title_font_family='Arial',
                                            hoverlabel_bgcolor='#DAEEED',   #Change the hover tooltip background color to a universal light blue color. If not specified, the background color will vary by team or completion pct, depending on what view the user chooses
                                            bargap=0.2,
                                            height=600,              
                                            xaxis_title="Tijd",
                                            yaxis_title="Omloopnummer",                  
                                            title_x=0.45,                    #Make title centered                    
                                            xaxis=dict()
                                        )
        
                    fig.update_yaxes(autorange='reversed')
                    fig.update_xaxes(tickformat='%H:%M')
                    st.plotly_chart(fig, use_container_width=True)
            
    else:
        st.warning("You need to upload a Excel file.")
        
# tabblad 3 ---------------------------------------------------------------------------------------------------------------------------------------------

with tab3:
    st.write('The routes of bus lines 400 and 401. De routes van buslijn 400 en 401. _(In the year 2022)_')
    image = Image.open('Routebussen.png')
    st.image(image, caption='Route of bus lines 400 en 401')

# tabblad 4 ---------------------------------------------------------------------------------------------------------------------------------------------

with tab4:
    st.write('VDL and Hermes')
    st_player("https://youtu.be/UD05b-Pkc8s")

# tabblad 5 ---------------------------------------------------------------------------------------------------------------------------------------------

with tab5:
    st.write('Transdev')
    st_player("https://youtu.be/Y3M1d5ML20Y")

# tabblad 6 ---------------------------------------------------------------------------------------------------------------------------------------------

with tab6:
    st.write("This video will show the user manual")
    st_player("https://youtu.be/IMWZQ3IXZFo")
    



    